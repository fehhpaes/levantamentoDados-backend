"""
Data Export Service

Export data in various formats: CSV, Excel, PDF, JSON
"""

import io
import json
from typing import Dict, List, Optional, Any, Union
from dataclasses import dataclass
from datetime import datetime
from enum import Enum
import logging

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    pd = None

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExportFormat(Enum):
    CSV = "csv"
    EXCEL = "excel"
    JSON = "json"
    PDF = "pdf"


@dataclass
class ExportConfig:
    """Configuration for data export."""
    format: ExportFormat
    filename: str
    title: Optional[str] = None
    include_metadata: bool = True
    include_charts: bool = False
    date_format: str = "%Y-%m-%d %H:%M:%S"
    decimal_places: int = 2


class DataExporter:
    """
    Multi-format data export service.
    
    Features:
    - CSV export with proper formatting
    - Excel export with styling and charts
    - PDF report generation
    - JSON export for APIs
    """
    
    def __init__(self):
        """Initialize data exporter."""
        self.available_formats = [ExportFormat.JSON]
        
        if PANDAS_AVAILABLE:
            self.available_formats.append(ExportFormat.CSV)
        
        if OPENPYXL_AVAILABLE:
            self.available_formats.append(ExportFormat.EXCEL)
        
        if REPORTLAB_AVAILABLE:
            self.available_formats.append(ExportFormat.PDF)
    
    def export(
        self,
        data: Union[List[Dict], Dict],
        config: ExportConfig
    ) -> bytes:
        """
        Export data to specified format.
        
        Args:
            data: Data to export
            config: Export configuration
            
        Returns:
            Bytes of exported file
        """
        if config.format not in self.available_formats:
            raise ValueError(
                f"Format {config.format.value} not available. "
                f"Available: {[f.value for f in self.available_formats]}"
            )
        
        if config.format == ExportFormat.CSV:
            return self._export_csv(data, config)
        elif config.format == ExportFormat.EXCEL:
            return self._export_excel(data, config)
        elif config.format == ExportFormat.JSON:
            return self._export_json(data, config)
        elif config.format == ExportFormat.PDF:
            return self._export_pdf(data, config)
    
    def _export_csv(
        self,
        data: Union[List[Dict], Dict],
        config: ExportConfig
    ) -> bytes:
        """Export to CSV format."""
        if not PANDAS_AVAILABLE:
            raise ImportError("pandas required for CSV export")
        
        if isinstance(data, dict):
            # Convert dict to list of records
            data = [data]
        
        df = pd.DataFrame(data)
        
        # Format dates
        for col in df.columns:
            if df[col].dtype == 'datetime64[ns]':
                df[col] = df[col].dt.strftime(config.date_format)
        
        # Round floats
        for col in df.select_dtypes(include=['float64']).columns:
            df[col] = df[col].round(config.decimal_places)
        
        output = io.StringIO()
        df.to_csv(output, index=False)
        
        return output.getvalue().encode('utf-8')
    
    def _export_excel(
        self,
        data: Union[List[Dict], Dict],
        config: ExportConfig
    ) -> bytes:
        """Export to Excel format with styling."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required for Excel export")
        
        if isinstance(data, dict):
            data = [data]
        
        wb = Workbook()
        ws = wb.active
        ws.title = config.title or "Data"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add metadata if requested
        if config.include_metadata:
            ws.append(["Generated:", datetime.now().strftime(config.date_format)])
            ws.append(["Records:", len(data)])
            ws.append([])
        
        if not data:
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()
        
        # Headers
        headers = list(data[0].keys())
        ws.append(headers)
        
        header_row = ws.max_row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        for row_data in data:
            row_values = []
            for header in headers:
                value = row_data.get(header, "")
                
                # Format values
                if isinstance(value, datetime):
                    value = value.strftime(config.date_format)
                elif isinstance(value, float):
                    value = round(value, config.decimal_places)
                elif isinstance(value, (dict, list)):
                    value = json.dumps(value)
                
                row_values.append(value)
            
            ws.append(row_values)
        
        # Apply borders to data
        for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[chr(64 + col_idx)].width = min(max_length + 2, 50)
        
        # Add charts if requested and applicable
        if config.include_charts and len(data) > 1:
            self._add_excel_charts(wb, ws, headers, header_row, len(data))
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def _add_excel_charts(
        self,
        wb: Any,
        ws: Any,
        headers: List[str],
        header_row: int,
        data_count: int
    ) -> None:
        """Add charts to Excel workbook."""
        # Find numeric columns for charts
        numeric_cols = []
        for idx, header in enumerate(headers, 1):
            try:
                cell = ws.cell(row=header_row + 1, column=idx)
                if isinstance(cell.value, (int, float)):
                    numeric_cols.append((idx, header))
            except:
                pass
        
        if len(numeric_cols) >= 2:
            # Create bar chart
            chart = BarChart()
            chart.title = "Data Overview"
            chart.style = 10
            
            data_ref = Reference(ws, min_col=numeric_cols[0][0], min_row=header_row,
                               max_row=header_row + data_count, max_col=numeric_cols[0][0])
            chart.add_data(data_ref, titles_from_data=True)
            
            ws.add_chart(chart, f"{chr(65 + len(headers) + 2)}2")
    
    def _export_json(
        self,
        data: Union[List[Dict], Dict],
        config: ExportConfig
    ) -> bytes:
        """Export to JSON format."""
        output = {
            "data": data
        }
        
        if config.include_metadata:
            output["metadata"] = {
                "exported_at": datetime.now().strftime(config.date_format),
                "record_count": len(data) if isinstance(data, list) else 1,
                "title": config.title
            }
        
        def json_serializer(obj):
            if isinstance(obj, datetime):
                return obj.strftime(config.date_format)
            raise TypeError(f"Type {type(obj)} not serializable")
        
        return json.dumps(output, indent=2, default=json_serializer).encode('utf-8')
    
    def _export_pdf(
        self,
        data: Union[List[Dict], Dict],
        config: ExportConfig
    ) -> bytes:
        """Export to PDF format."""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab required for PDF export")
        
        if isinstance(data, dict):
            data = [data]
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30
        )
        
        # Title
        if config.title:
            elements.append(Paragraph(config.title, title_style))
        
        # Metadata
        if config.include_metadata:
            meta_style = styles['Normal']
            elements.append(Paragraph(
                f"Generated: {datetime.now().strftime(config.date_format)}",
                meta_style
            ))
            elements.append(Paragraph(f"Records: {len(data)}", meta_style))
            elements.append(Spacer(1, 20))
        
        if not data:
            doc.build(elements)
            return output.getvalue()
        
        # Table
        headers = list(data[0].keys())
        table_data = [headers]
        
        for row_data in data:
            row = []
            for header in headers:
                value = row_data.get(header, "")
                
                if isinstance(value, datetime):
                    value = value.strftime(config.date_format)
                elif isinstance(value, float):
                    value = round(value, config.decimal_places)
                elif isinstance(value, (dict, list)):
                    value = str(value)[:50]  # Truncate complex data
                
                row.append(str(value))
            table_data.append(row)
        
        # Calculate column widths
        col_width = 500 / len(headers)
        
        table = Table(table_data, colWidths=[col_width] * len(headers))
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7fafc')])
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        return output.getvalue()
    
    # Specialized export methods
    def export_predictions(
        self,
        predictions: List[Dict],
        format: ExportFormat,
        filename: str = "predictions"
    ) -> bytes:
        """Export match predictions."""
        config = ExportConfig(
            format=format,
            filename=filename,
            title="Match Predictions Report",
            include_metadata=True
        )
        
        return self.export(predictions, config)
    
    def export_value_bets(
        self,
        value_bets: List[Dict],
        format: ExportFormat,
        filename: str = "value_bets"
    ) -> bytes:
        """Export value bet opportunities."""
        config = ExportConfig(
            format=format,
            filename=filename,
            title="Value Bet Opportunities",
            include_metadata=True,
            decimal_places=4
        )
        
        return self.export(value_bets, config)
    
    def export_backtest_results(
        self,
        results: Dict,
        format: ExportFormat,
        filename: str = "backtest_results"
    ) -> bytes:
        """Export backtest results."""
        # Flatten results for export
        flattened = {
            "strategy": results.get("strategy_name"),
            "period_start": results.get("period_start"),
            "period_end": results.get("period_end"),
            "total_bets": results.get("total_bets"),
            "win_rate": results.get("win_rate"),
            "roi": results.get("roi"),
            "max_drawdown": results.get("max_drawdown"),
            "sharpe_ratio": results.get("sharpe_ratio"),
            "profit_factor": results.get("profit_factor")
        }
        
        config = ExportConfig(
            format=format,
            filename=filename,
            title="Backtest Results",
            include_metadata=True,
            include_charts=True
        )
        
        return self.export([flattened], config)
    
    def export_bankroll_history(
        self,
        history: Dict,
        format: ExportFormat,
        filename: str = "bankroll_history"
    ) -> bytes:
        """Export bankroll transaction history."""
        transactions = history.get("transactions", [])
        
        config = ExportConfig(
            format=format,
            filename=filename,
            title="Bankroll History",
            include_metadata=True
        )
        
        return self.export(transactions, config)
    
    def get_available_formats(self) -> List[str]:
        """Get list of available export formats."""
        return [f.value for f in self.available_formats]
    
    def get_mime_type(self, format: ExportFormat) -> str:
        """Get MIME type for export format."""
        mime_types = {
            ExportFormat.CSV: "text/csv",
            ExportFormat.EXCEL: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ExportFormat.JSON: "application/json",
            ExportFormat.PDF: "application/pdf"
        }
        return mime_types.get(format, "application/octet-stream")
    
    def get_file_extension(self, format: ExportFormat) -> str:
        """Get file extension for export format."""
        extensions = {
            ExportFormat.CSV: ".csv",
            ExportFormat.EXCEL: ".xlsx",
            ExportFormat.JSON: ".json",
            ExportFormat.PDF: ".pdf"
        }
        return extensions.get(format, ".bin")
